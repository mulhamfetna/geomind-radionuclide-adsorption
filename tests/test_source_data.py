import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from geomind import source_data as S  # noqa: E402


def test_source_data_has_a_sheet_per_figure_and_the_pools():
    d = S.build_source_data()
    expected = {"Fig2a_framework_Al", "Fig2b_surface_area", "Fig2c_causal_test",
                "Fig3_forward_model", "Fig4_structural", "Fig5_saturation",
                "Fig6_pooling_limit", "PoolA_adsorption", "PoolB_immobilisation"}
    assert expected <= set(d)
    for name, df in d.items():
        assert len(df) > 0, f"{name} is empty"


def test_source_data_matches_the_published_numbers():
    d = S.build_source_data()
    # the forward model is fit on 7 Varon samples
    assert len(d["Fig3_forward_model"]) == 7
    # the causal test matches all 7 fresh<->leached pairs (F27)
    assert len(d["Fig2c_causal_test"]) == 7
    # the audited pools
    assert len(d["PoolA_adsorption"]) == 141
    assert len(d["PoolB_immobilisation"]) == 73
    # pooled Cs Langmuir set behind Fig 6b
    assert len(d["Fig6_pooling_limit"]) == 23


def test_write_xlsx_produces_one_workbook_with_every_sheet(tmp_path):
    out = S.write_xlsx(tmp_path / "source-data.xlsx")
    assert Path(out).exists()
    import openpyxl
    wb = openpyxl.load_workbook(out)
    assert "Fig3_forward_model" in wb.sheetnames
    assert "README" in wb.sheetnames          # a sheet explaining provenance + licence


def test_complete_database_export_includes_the_excluded_rows():
    """The export must document what was removed, not only what was kept."""
    from geomind import database_export as DB
    s = DB.build_sheets()
    assert len(s["01 Pool A adsorption"]) == 141
    assert len(s["02 Pool B immobilisation"]) == 73
    assert len(s["03 Excluded rows"]) >= 8
    assert set(s["03 Excluded rows"]["action"]) <= {"REMOVED", "RELABELLED", "RE-TYPED"}
    assert (s["03 Excluded rows"]["reason"].str.len() > 20).all()
    assert len(s["04 Audit trail"]) == 91
    assert len(s["05 Findings"]) == 46
    assert len(s["12 Meta not admitted"]) >= 7


def test_dataset_deposit_bundle_is_complete_and_self_describing():
    from geomind import dataset_deposit as DD
    DD.build()
    d = DD.OUT_DIR
    for f in ["GEOMIND-R-complete-database.xlsx", "GEOMIND-R-source-data.xlsx",
              "pool_a_adsorption.csv", "pool_b_immobilisation.csv",
              "DATA-DICTIONARY.md", "README.md", "CITATION.cff", "LICENSE"]:
        assert (d / f).exists(), f"missing {f}"
    import pandas as pd
    assert len(pd.read_csv(d / "pool_a_adsorption.csv")) == 141
    assert len(pd.read_csv(d / "pool_b_immobilisation.csv")) == 73
    dd = (d / "DATA-DICTIONARY.md").read_text()
    assert "al_iv_mmol_g" in dd and "ca_si_al" in dd      # the key descriptors documented
    assert "CC BY 4.0" in (d / "README.md").read_text()
    assert "type: dataset" in (d / "CITATION.cff").read_text()
