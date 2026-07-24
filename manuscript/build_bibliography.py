"""Build the corpus bibliography deliverables: an Excel table and a Vancouver-style DOCX.

Both list every source publication this study drew on, with authoritative metadata resolved
from Crossref by DOI, the publisher's own abstract where it is available, and an Arabic
translation of that abstract.

    python -m build_bibliography      # writes GEOMIND-R-bibliography.{xlsx,docx}

## How the abstracts were sourced, and why four are absent

Every abstract is verbatim and attributable. Each comes from one of two sources, recorded per
row in ``abstract_source``: the publisher's own text supplied through Crossref, or the text
read directly from the paper's PDF and **verified by eye against that PDF** (title checked,
column-aware extraction to strip keyword sidebars, de-hyphenated only). Naive full-text
extraction was explicitly *not* trusted — an early automated pass produced a confirmed
wrong-paper match (a slag/Portland-cement abstract attached to a zeolite ion-exchange paper),
so every extracted abstract was located and confirmed individually.

Four papers carry **no** abstract and resolve to their DOI instead, for stated reasons rather
than omission: two are not held on disk in a readable form (``katada2024`` absent; ``qian2001``
a scanned image with no text layer), and two have a two-column first page whose keyword sidebar
is interleaved into the abstract's own text band and cannot be separated without reconstructing
words — which would risk fabrication (``jain2022``, ``niu2022``). Their key results are still
recorded verbatim in the finding register and the source data.

## Copyright

Publisher abstracts are the copyright of their publishers. This compilation is prepared as a
working bibliography for the authors' own submission; it is **not** part of the public data
deposit. Each row records the licence Crossref reports, where one is declared.
"""
from __future__ import annotations

import json
import subprocess
import sys
from pathlib import Path

_HERE = Path(__file__).resolve().parent
_ROOT = _HERE.parent
#: Committed source metadata: Crossref bibliographic fields per DOI, plus abstracts. The
#: 28 abstracts present were each verified against the paper's own PDF or supplied verbatim
#: by the publisher via Crossref (see abstract_source). Two (katada2024, qian2001) carry no
#: abstract — the first is not held on disk, the second is a scanned PDF with no text layer —
#: and resolve to their DOI instead.
SOURCES = _HERE / "bibliography_sources.json"

#: What each source contributed, so a reader can see why it is in the corpus at all.
ROLE = {
    "geomind": "Reference work this project set out to replicate",
    "varon2025": "Forward-model training series (n=7): framework Al-IV vs Sr K_D",
    "varon_leached2026": "Within-sample causal test: leaching strips Al-IV and uptake falls",
    "oulu2026": "Designed Si/Al series with published Q4(mAl) deconvolution; ARI source",
    "walkley2020": "Atomic mechanism: Sr/Ca displace Na/K from Al-IV charge-balancing sites",
    "geddes2025": "Atomic mechanism: Sr imaged at Al-IV sites (EXAFS/NMR)",
    "vandevenne2018": "Designed Ca-Si-Al slag series; quantifies the Ca/Sr site competition",
    "jain2022": "Cs loading series; pollucite route to retention",
    "blackford2007": "TEM/NMR: Sr partitions to crystalline SrCO3 at waste loadings",
    "nevin2026": "Pool B: Cs/Sr leachability indices with NMR structure",
    "kim2026": "Pool B: Sr leachability with EXAFS coordination numbers",
    "jang2016": "Pool B: Cs and Sr leachability with pore structure",
    "kurumisawa2021": "Pool B: Cs ingress diffusivity",
    "komljenovic2020": "Pool B: Ca-system (C-A-S-H), quantitative confirmation of F19",
    "arbelhaddad2022": "Pool B: cumulative leached fraction, metakaolin framework",
    "stanojevic2025": "Pool B: Cs leachability, fly-ash geopolymer",
    "frederickx2025": "Pool B: real waste stream; accelerated-leachant caution",
    "katada2024": "Pool A: Na-zeolite Cs ion-exchange capacities",
    "baek2018": "Pool A: natural zeolites; source of a saturation-screen artefact",
    "tarnovsky2024": "Pool A: metakaolin geopolymers, Cs and Sr with BET",
    "xiang2021": "Pool A: geopolymer foams; BET falls while uptake rises",
    "zhang2021": "Pool A: alkali-activated fly-ash/slag geomaterials, Cs",
    "zheng2023": "Pool A: zeolite-rich geopolymer, Cs",
    "lei2021": "Pool A: Cs Freundlich / Sr Langmuir",
    "elnaggar2018": "Pool A: MKBFS systems (two rows quarantined by the audit)",
    "niu2022": "Pool A: ion-exchange log K only; no capacities (audit outcome)",
    "tian2019_espr": "Pool A: zeolite A from fly ash",
    "tian2019_wst": "Pool A: fly-ash geopolymer, Cs/Sr; values in the supplement",
    "lin2026": "Pool A: geopolymer granules, Cs/Sr in simulated seawater",
    "qian2001": "Pool A: Sr/Cs distribution coefficients",
    "petlitckaia2020": "Pool A: geopolymer foams, Cs",
    "nadol2026": "Context: base-element leaching + NMR; no radionuclide target",
    "hamed2025_magnetic": "Pool A: magnetic nanocomposite (audit correction)",
    "muracchioli2019": "Rejected: heavy-metal adsorption, out of scope",
    "weigelt2020": "Context: structural characterisation of Cs/Sr geopolymers",
    "perera2004": "Context: no primary quantitative leach data",
}

#: Arabic translations of the publisher abstracts. Authored for this compilation; the English
#: column remains the authoritative verbatim text.
ARABIC: dict[str, str] = {
 "blackford2007":
  "أظهر الفحص بالمجهر الإلكتروني النافذ لطور جيوبوليمر مُحضَّر من الميتاكاولين ومحاليل السيليكات "
  "القلوية، بنسب موليّة اسمية Na/Al تساوي 1 وSi/Al تساوي 2، أنه غير متبلور على مقياس نحو 1 نانومتر "
  "بعد المعالجة عند 40 °م. وفي العينات المحتوية على 5% وزناً من السيزيوم أو السترونتيوم، استقر "
  "السيزيوم في الطور غير المتبلور، بينما أُدمج السترونتيوم جزئياً فقط إذ فُصل تفضيلياً إلى كربونات "
  "السترونتيوم SrCO₃ المتبلورة. ولم يكن للتسخين التدريجي لنزع الماء أثر يُذكر في البنية الكلية حتى "
  "500 °م. وبيّن الرنين النووي المغناطيسي في الحالة الصلبة، للمادة المعالَجة قرب درجة حرارة الغرفة، "
  "أن السيزيوم — شأنه شأن الصوديوم — يرتبط أساساً بماء المسام، لكن مع ارتباط ملحوظ بالشبكة "
  "الألومينوسيليكاتية أكبر مما هو عليه لدى الصوديوم. وقد زاد التسخين اللاحق حتى 300 °م من ارتباط "
  "السيزيوم والصوديوم بالشبكة.",
 "frederickx2025":
  "في سياق التخلص من الوقود النووي المستهلك، تمثل النويدات المشعة المولّدة للحرارة مثل السيزيوم "
  "والسترونتيوم مصدر قلق بالغ، لما لها من تأثير كبير في المسافة الواجب تركها بين أنفاق التخلص، "
  "ومن ثمّ في كلفة منشأة التخلص. لذلك تبحث بعض السيناريوهات في فصل الوقود المستهلك وتحويله لتحسين "
  "قابلية التخلص من مجاري النفايات الغنية بالسيزيوم والسترونتيوم وما تبقّى منها. وفي هذه الدراسة "
  "جرى تثبيت مجرى نفايات غني بالسيزيوم والسترونتيوم، وهو محلول نتراتي، ضمن مصفوفات منشَّطة قلوياً "
  "قائمة على الميتاكاولين وخبث الأفران العالية. وقد اختيرت هذه المصفوفات للتثبيت لما هو معروف عنها "
  "من مزايا في المتانة و/أو مقاومة الحرارة مقارنةً بالمواد الإسمنتية التقليدية. وهدف الدراسة هو "
  "تطوير خلطة مثلى لاحتجاز السيزيوم والسترونتيوم.",
 "geomind":
  "‏GEOMIND نموذج تعلّم آلي هجين دُرِّب على قاعدة بيانات داخلية للتوصية بتركيبات جيوبوليمرية مثلى "
  "لخصائص مستهدفة يحددها المستخدم.",
 "lin2026":
  "لا تزال الإزالة الانتقائية للسيزيوم-137 والسترونتيوم-90 المشعّين من مياه الصرف المشعة عالية "
  "الملوحة تحدياً جوهرياً، إذ تخفض الأيونات المنافسة كفاءة الامتزاز وانتقائيته. طُوِّرت في هذه "
  "الدراسة مواد مازّة محبَّبة عالية الأداء قائمة على مصفوفات جيوبوليمرية منشَّطة قلوياً لتحسين أداء "
  "الامتزاز. وقد رُكِّبت المازّات بالبلمرة غير العضوية، وأمكن الحصول على حبيبات متينة ميكانيكياً "
  "ذات مسامية وكيمياء سطحية مضبوطتين. وأظهرت تجارب الامتزاز الدفعية في ماء بحر محاكى كفاءات إزالة "
  "تتجاوز 99% للسيزيوم والسترونتيوم، وأكدت نمذجة الأيزوثرم سعات امتزاز عظمى مرتفعة (حتى 0.41 "
  "مكافئ ميلي/غ للسيزيوم و5.07 مكافئ ميلي/غ للسترونتيوم). كما بيّنت اختبارات الأعمدة ذات الحشوة "
  "الثابتة استمرار كفاءات الإزالة للمازّات المُحسَّنة.",
 "nevin2026":
  "تُعدّ الجيوبوليمرات بديلاً واعداً لأشكال النفايات التقليدية القائمة على الإسمنت البورتلاندي في "
  "تثبيت نواتج الانشطار المشعة الخطرة مثل السيزيوم-137 والسترونتيوم-90، إذ توفر متانة أعلى ومعدلات "
  "ترشيح أدنى.",
 "tarnovsky2024":
  "هدف العمل المقدَّم إلى تخليق جيوبوليمرات قائمة على الميتاكاولين وتحديد قدرتها على الامتزاز في "
  "عملية إزالة أيونات السيزيوم والسترونتيوم من المحاليل المائية. واقتُرحت مقاربات جديدة للحصول على "
  "عيّنتين من الجيوبوليمرات بصور ملائمة تقنياً. ودُرست مورفولوجيا المواد بوساطة تحليل الفلورة "
  "السينية (XRF)، وامتزاز/انفكاك النيتروجين عند درجات حرارة منخفضة، والمجهر الإلكتروني الماسح "
  "(SEM). وتبيّن من تحليل XRF أن أكسيدَي SiO₂ وAl₂O₃ يمثلان المكوّنين الرئيسين في جميع العينات "
  "المدروسة (نحو 54–84% وزناً). وأظهرت دراسات SEM أن الجيوبوليمرات تتألف من جسيمات نانوية ورابط "
  "جيوبوليمري غير متبلور وكاولين غير متفاعل، وأن جميع العينات تحوي مسامات متوسطة بأنصاف أقطار نحو "
  "1–40 نانومتر.",
 "tian2019_wst":
  "البلمرة الجيولوجية عملية تفاعلية آخذة في التطور للاستفادة من النفايات الصلبة. في هذه الدراسة "
  "جرى تخليق جيوبوليمر قائم على الرماد المتطاير ومشتقّه (جيوبوليمر معدَّل بالحديد الثنائي) "
  "وتوصيفهما بوساطة XRD وSEM وFTIR وBET وUV-Vis DRS وTG-DTA، واستُخدما مازّين لإزالة أيونات "
  "السيزيوم والسترونتيوم من المحاليل. وقد توافقت حركية الامتزاز جيداً مع نموذج الرتبة الثانية "
  "الزائفة، وتوافق امتزاز السيزيوم والسترونتيوم على الجيوبوليمر الأصلي بصورة أفضل مع نموذج لانغموير، "
  "في حين كان نموذج فرويندليخ أنسب للامتزاز على الجيوبوليمر المعدَّل بالحديد الثنائي. وأشارت "
  "الطاقات الحرة المحسوبة من أيزوثرم D-R إلى أن الامتزاز يجري أساساً بالتبادل الأيوني، وأن حجم "
  "الحلقة يؤدي دوراً حاسماً في التبادل الأيوني لكلا الأيونين، كما أن لترتيب رباعيات السطوح SiO₄ "
  "وAlO₄ أثراً مهماً.",
 "zhang2021":
  "دُرست في هذا البحث إزالة السيزيوم من المحاليل المائية باستخدام مواد جيولوجية، إذ اختير الامتزاز "
  "طريقةً فعّالة لتطويرها لإزالة السيزيوم من سوائل النفايات المشعة. وحُضِّرت المواد الجيولوجية، "
  "ومنها الرماد المتطاير والخبث كمواد أولية، بوصفها مازّات باستخدام منشِّط قلوي. وجرى توصيف المواد "
  "بحيود الأشعة السينية (XRD)، والمجهر الإلكتروني الماسح المزوَّد بمطياف تشتت الطاقة (SEM-EDS)، "
  "وتحليل المساحة السطحية بطريقة BET وحجم المسام وقياسها. ودُرس تأثير عوامل مختلفة مثل الأس "
  "الهيدروجيني وزمن التماس وجرعة المازّ في امتزاز السيزيوم. وقُيِّم معامل التوزيع والسعة الامتزازية "
  "لتقدير الأداء الفعلي للمازّ، وأظهرت المواد القائمة على الرماد المتطاير سعة امتزاز عظمى للسيزيوم "
  "بلغت 89.32 ملغ/غ ومعامل توزيع مرتفعاً قدره 31.02 ملغ·غ⁻¹·ميلي مول⁻¹.",
 "zheng2023":
  "يولّد تشغيل محطات الطاقة النووية كميات كبيرة من سوائل النفايات المشعة المنخفضة والمتوسطة "
  "المستوى. ويمكن للجيوبوليمرات الغنية بالزيوليت، المخلَّقة في ظروف حرارية مائية من الرماد المتطاير "
  "الصناعي، أن تثبّت النويدات المشعة بفعالية. وقد دُرس في هذا البحث قانون تخليق الجيوبوليمرات "
  "الغنية بالزيوليت وأداء امتزاز/انفكاك النويدة المشعة ⁺Cs باستخدام XRD وSEM وICP. وتُظهر النتائج "
  "أن زيادة درجات حرارة المعالجة وتراكيز هيدروكسيد الصوديوم تؤدي إلى تحوّل الزيوليت من النمط Y إلى "
  "الشابازيت والكانكرينيت عند تراكيز منخفضة من نترات الصوديوم، في حين لا يكون لهيدروكسيد الصوديوم "
  "فوق 2 مول تأثير واضح في التحوّل الطوري عند التراكيز المرتفعة.",
 "baek2018":
  "دُرست خصائص امتزاز السيزيوم والتبادل الأيوني التنافسي مع كاتيونات قلوية أخرى ومع السترونتيوم على "
  "ثلاثة زيوليتات طبيعية هي الشابازيت والستيلبيت والهيولانديت. أظهر الشابازيت أسرع امتزاز للسيزيوم، "
  "يليه الهيولانديت ثم الستيلبيت، وارتبط ذلك ارتباطاً وثيقاً بسعته العالية للتبادل الكاتيوني (CEC) "
  "ونسبة Si/Al المنخفضة. وقد فُسِّرت البيانات الحركية لجميع الزيوليتات على أفضل نحو بنموذج الرتبة "
  "الثانية الزائفة، وتضمّنت عملية الامتزاز خطوات متعددة تتأثر بقيم CEC. وتوافقت أيزوثرمات الامتزاز "
  "التوازنية للزيوليتات الثلاثة مع نموذجَي لانغموير وفرويندليخ معاً. وأظهرت أيزوثرمات التبادل أن "
  "للشابازيت انتقائية للسيزيوم أعلى من غيره؛ وعند الكسور المكافئة المنخفضة للسيزيوم في المحلول كان "
  "ترتيب الانتقائية للكاتيونات الأخرى: Na > Li > Sr > K > Rb. أمّا في الستيلبيت والهيولانديت فتكون "
  "انتقائية السيزيوم أعلى فقط عند الكسور المكافئة المنخفضة، وتتّخذ الأيزوثرمات أشكالاً سينية لأن "
  "للسيزيوم أكثر من موقعَي امتزاز فيهما، بينما للشابازيت موقع واحد.",
 "elnaggar2018":
  "يشكّل السيزيوم المشعّ تهديدات بيئية كبيرة. وامتزاز الأنواع الخطرة على مازّات جيوبوليمرية أمرٌ حديث "
  "نسبياً قد يفيد في فهم آليات الاحتجاز عند استخدام الجيوبوليمرات لتثبيت النفايات المشعة. هنا خُلِّقت "
  "مازّات جيوبوليمرية Na-MK وK-MK وNa-MKBFS وK-MKBFS من الميتاكاولين وخبث الأفران العالية، ووُصِّفت "
  "بحيود الأشعة السينية والفلورة السينية والأشعة تحت الحمراء والتحليل الحراري والمجهر الإلكتروني "
  "الماسح. وأوضحت نتائج FT-IR/XRF أثر الكاتيون القلوي أحادي التكافؤ (M⁺) في تقسيم المازّات إلى "
  "غنية بالألمنيوم (صوديومية) وغنية بالسيليكون (بوتاسيومية). وكانت جميع المازّات غير متبلورة إلى "
  "شبه متبلورة. وأُنشئت أيزوثرمات امتزاز النويدة ¹³⁴Cs بتأثير حراري موجب. وأعطت المازّات الغنية "
  "بالألمنيوم سعات امتزاز أعلى من الغنية بالسيليكون، إذ سجّل المازّ Na-MK أميز سعة امتزاز "
  "(74.95 ملغ/غ عند 333 كلفن). واستُخدمت نماذج لانغموير وفرويندليخ وD-R للكشف عن السعات والآليات. "
  "وكان امتزاز ⁺Cs مواتياً على جميع المازّات، وحكمَته آلية التبادل الأيوني في كل النظم عدا نظام "
  "¹³⁴Cs/K-MK الذي حكمَته آلية الامتزاز الفيزيائي.",
 "jang2016":
  "تبحث هذه الدراسة الأثر الحاجزي الفيزيائي لأشكال النفايات الجيوبوليمرية في سلوك ترشيح السيزيوم "
  "والسترونتيوم. استُخدمت جيوبوليمرات قائمة على الرماد المتطاير وأخرى ممزوجة بالخبث عواملَ تصليب، "
  "وقُيِّم سلوك الترشيح وفق المعيار ANSI/ANS-16.1. وكانت انتشارية السيزيوم والسترونتيوم في "
  "الجيوبوليمر القائم على الرماد المتطاير أدنى مما في الإسمنت البورتلاندي بعاملَي 10³ و10⁴ على "
  "التوالي، مبيّنةً أداء تثبيت محسَّناً جوهرياً. وكانت مقاومة الترشيح ثابتة نسبياً بصرف النظر عن نوع "
  "الرماد المتطاير. وارتبطت انتشارية أيونات السيزيوم والسترونتيوم الذائبة في الماء ارتباطاً وثيقاً "
  "بقطر المسام الحرج للرابط. وكان قطر المسام الحرج للجيوبوليمر القائم على الرماد المتطاير أصغر بكثير "
  "منه في الإسمنت البورتلاندي والجيوبوليمر الممزوج بالخبث؛ ومن ثمّ كانت قدرته على إعاقة انتشار "
  "النويدات فيزيائياً (الأثر الحاجزي الفيزيائي) متفوّقة.",
 "kim2026":
  "تُعدّ الإسمنتات الجيوبوليمرية موادّ واعدة جداً للتثبيت طويل الأمد لنفايات السترونتيوم-90 المشعة، "
  "لما توفّره من متانة فائقة ومواقع ربط كاتيونية مقارنةً بمصفوفات الإسمنت البورتلاندي التقليدية. "
  "تبحث هذه الدراسة أثر الترشيح المطوَّل في آلية تثبيت السترونتيوم والسلامة البنيوية للجيوبوليمرات "
  "القائمة على الميتاكاولين باستخدام اختبار الترشيح شبه الديناميكي ANSI/ANS 16.1. أبدت جميع "
  "الجيوبوليمرات احتجازاً عالياً للسترونتيوم بمعامل قابلية ترشيح لا يقلّ عن 14.7 لكل العينات، متجاوزاً "
  "بكثير الإرشاد الصناعي البالغ 6.0. والأهمّ أن الجيوبوليمرات المنشَّطة بسيليكات البوتاسيوم أبدت "
  "إطلاقاً أقلّ للسترونتيوم ومعدّلات ترشيح أدنى بكثير من نظيرتها المنشَّطة بسيليكات الصوديوم. وكشف "
  "التحليل الطيفي والحيودي متعدّد المقاييس — بما فيه مطيافية امتصاص الأشعة السينية السنكروترونية "
  "والرنين النووي المغناطيسي في الحالة الصلبة عالي الحقل للنوى ³⁹K و²³Na و²⁷Al و²⁹Si — أن شبكة "
  "هلام الألومينوسيليكات القلوية بقيت مستقرة بنيوياً بعد الترشيح 28 يوماً دون تغيّرات تُذكر في بيئات "
  "روابط السيليكون والألمنيوم. ويُتحكَّم بإطلاق السترونتيوم أساساً بالانتشار، والآلية السائدة للتثبيت "
  "هي تكوّن كربونات السترونتيوم غير الذائبة.",
 "komljenovic2020":
  "تُستخدم الروابط المنشَّطة قلوياً (AABs)، كبديل واعد للإسمنت البورتلاندي، على نطاق تجاري في تطبيقات "
  "شتّى حول العالم، منها تثبيت النفايات الخطرة والمشعة. وفي هذه الورقة دُرست مقاومة الترشيح والمتانة "
  "والتغيّر النانوي البنيوي لخبث الأفران العالية المنشَّط قلوياً (AABFS) المطعَّم بنسبتَي 2% و5% من "
  "السيزيوم. وأدّت إضافة السيزيوم إلى زيادة معتبرة في مقاومة الانضغاط، أعقبها انخفاض طفيف بعد الترشيح. "
  "ويمكن عدّ AABFS مصفوفةً كفؤةً محتملةً لتثبيت السيزيوم، إذ كان متوسط معامل قابلية الترشيح في "
  "الحالتين (2% و5%) فوق القيمة الحدّية 6. وقد سبّب كلٌّ من التطعيم بالسيزيوم والترشيح تحوّلاً في البنية "
  "النانوية لِـ AABFS: فمعظم الألمنيوم المُطلَق من هلام C-A-S-H بفعل الترشيح بقي داخل المصفوفة "
  "مُطلِقاً إعادة بناء الهلام؛ إذ تحوّل هلام C-A-S-H إلى C-S-H وتكوّن هلام إضافي N-(C)-A-S-H. وارتبط "
  "السيزيوم تفضيلياً بهلام N-(C)-A-S-H بدل C-A-S-H. وتبدو النتائج متوافقة مع نموذج التوبرموريت "
  "المستبدَل المتشابك (CSTM).",
 "kurumisawa2021":
  "أُجري في هذه الدراسة تحقيق أساسي فيما إذا كان يمكن استخدام الجيوبوليمر جزءاً من حاجز اصطناعي أثناء "
  "التخلص من النفايات المشعة. تتألف الجيوبوليمرات أساساً من الألومينا والسيليكا، وتُبدي ترشيحاً "
  "مهمَلاً لغياب الكالسيوم. والدراسات عليها محدودة مقارنةً بالمواد الإسمنتية الأخرى لأن خصائصها "
  "الفيزيائية تتغيّر بتغيّر ظروف الإنتاج. في هذا العمل حُضِّرت جيوبوليمرات قائمة على الميتاكاولين "
  "وحُلِّل أداؤها الانتشاري. وتشير النتائج إلى أن انتشارية السيزيوم في الجيوبوليمر تتأثر بنوع المنشِّط "
  "القلوي؛ إذ كانت الجيوبوليمرات المنشَّطة بالصوديوم أعلى في سعة امتزاز السيزيوم من المنشَّطة "
  "بالبوتاسيوم، وكان لسعة امتزاز السيزيوم أثر معتبر أيضاً في انتشاريته داخل الجيوبوليمرات.",
 "lei2021":
  "نُبلِّغ هنا عن تصنيع كُريّات جيوبوليمرية دقيقة قائمة على الميتاكاولين/الخبث بتقنية التشتّت-التعليق-"
  "التصلّب، ثم تحويلها إلى كُريّات زيوليتية دقيقة بمعالجة حرارية موضعية. وقد تحسّنت الخصائص الريولوجية "
  "والمتانة الميكانيكية لكُريّات الزيوليت (M/SZMs) بإضافة الخبث، ووُصِّفت نسيجياً ومورفولوجياً بـ BET "
  "وSEM-EDX وXRD. وعند محتوى خبث 20% من الكتلة الكلية ازدادت المساحة السطحية النوعية زيادةً معتبرة "
  "دون تغيير بنية الزيوليت. وأظهر تحليل الخصائص الريولوجية للمعلّق طوراً مائعاً لدائنياً زائفاً يوافق "
  "نموذج هيرشل-بالكلي. وتبع امتزاز ⁺Cs و²⁺Sr من مياه الصرف حركية الرتبة الثانية الزائفة، وكانت "
  "السعة العظمى 103.74 ملغ/غ للسيزيوم و54.90 ملغ/غ للسترونتيوم، وفُسِّرتا على أفضل نحو بنموذجَي "
  "فرويندليخ ولانغموير على التوالي. وأبدت الكُريّات فصلاً ديناميكياً ممتازاً في نظام أعمدة، وأداءً "
  "بارزاً في الإزالة من عينات مياه صرف حقيقية مختلفة؛ ومع بساطة التصنيع وانخفاض الكلفة وارتفاع "
  "الكفاءة، يمكن عدّها مرشَّحاً بديلاً لإزالة السيزيوم والسترونتيوم من مياه الصرف.",
 "nadol2026":
  "تهدف هذه الدراسة إلى استكشاف كيفية تأثير الترشيح في البنية المجهرية لجيوبوليمر مختار مشتقّ من "
  "الرماد المتطاير، وتقييم إمكان استخدامه للتخلص السطحي القريب من النفايات المشعة. وتكمن فرادة هذه "
  "التركيبة في قدرتها على تكوين شبكة ألومينوسيليكاتية تبقى مقاومةً للتغيّرات البنيوية عند ملامسة "
  "الماء. وكان معامل ترشيح العناصر الأساسية مثل الكالسيوم والألمنيوم والصوديوم والسيليكون بين 11 و13، "
  "مما يوحي بتفوّق التركيبة على الإسمنت البورتلاندي أو الإسمنت الهيدروليكي التقليدي. ويكشف الرنين "
  "النووي المغناطيسي في الحالة الصلبة أن العينة المتفاعلة مع الماء خالية من تلوّث الطور الزيوليتي، "
  "وأن كل الصوديوم مُدمَج في شبكة هلام الجيوبوليمر الألومينوسيليكاتية، مؤكِّداً أن الشبكة الزجاجية "
  "للجيوبوليمر هي المسؤولة عن انخفاض قابلية ترشيح العناصر الأساسية، وأنه لا يوجد صوديوم متحرّك متاح "
  "للتبادل كما في الزيوليتات.",
 "oulu2026":
  "يؤثّر تركيب المواد المنشَّطة قلوياً (AAMs) والجيوبوليمرات في خصائصها المادية وأدائها في تطبيقات "
  "الامتزاز، ومع ذلك تبقى ارتباطاتها غير مستكشَفة إلى حدّ كبير. في هذه الدراسة خُلِّقت هذه المواد "
  "بتغيير تركيبها منهجياً في المديَين Si1Al1Na1–Si20Al1Na1 (أي جيوبوليمرات خالية من الكالسيوم) "
  "وSi1Al1Na1Ca2–Si20Al1Na1Ca21 (أي موادّ حاوية للكالسيوم). ورُبطت الخصائص المادية (ترابط البنية "
  "الألومينوسيليكاتية، والمساحة السطحية النوعية، وحجم المسام، ومتوسط حجمها، وجهد زيتا) بأداء الامتزاز "
  "لكاتيونات ذات أنصاف أقطار مائية مختلفة: أزرق الميثيلين (MB)، ورودامين 6G، والأمونيوم (⁺NH₄). ففي "
  "الجيوبوليمرات الخالية من الكالسيوم ازداد امتزاز MB وR6G بزيادة نسبة Si/Al وارتبط بقوة بالمساحة "
  "السطحية النوعية، بينما أبدى امتزاز ⁺NH₄ اتجاهاً معاكساً، مرتبطاً إيجاباً بنسبة Al/Si وجهد زيتا "
  "وسلباً بالمساحة السطحية. وأدّت إضافة الكالسيوم إلى بلوغ كميات الامتزاز حدّاً أدنى عند التركيبة "
  "Si5Al1Na1Ca6، في حين رفعها محتوى الكالسيوم الأدنى أو الأعلى.",
 "petlitckaia2020":
  "خُلِّقت رغوة جيوبوليمرية تحتضن شبكة ثلاثية الأبعاد من المسام المترابطة، ووُظِّفت بهكسا سيانو "
  "فيرات نحاس البوتاسيوم [K2CuFe(CN)6] بغية إزالة تلوّث النفايات السائلة المشعة الحاوية للسيزيوم. "
  "ووُصِّفت الرغوة الجيوبوليمرية (GF) والرغوات الموظَّفة (FGF) بمجموعة تقنيات (SEM وTEM وXRD …) قبل "
  "دراسة قدرتها على إزالة السيزيوم انتقائياً من المحلول. ويترسّب K2CuFe(CN)6 بانتظام على جدران مسام "
  "الرغوة وداخل الشبكة المتوسطة المسام. وباستثمار الحركية وأيزوثرمات الامتزاز في محاليل مختلفة (ماء "
  "منزوع الأيونات، وماء عذب، وماء عذب بفائض من الصوديوم) أُجريت دراسة مقارِنة بين GF وFGF. وفي وجود "
  "أيونات منافِسة في المحلول تنخفض سعة المادّتين مقارنةً بالماء منزوع الأيونات، وتكون آلية التبادل "
  "‏⁺Na ⇄ ⁺Cs و⁺K ⇄ ⁺Cs في GF وFGF على التوالي.",
 "stanojevic2025":
  "تبحث هذه الدراسة أثر إضافة 2% و5% وزناً من السيزيوم في زمن الشكّ ومقاومة الانضغاط وبنية "
  "الجيوبوليمرات القائمة على الرماد المتطاير، إضافةً إلى أثر ترشيح السيزيوم في ماء منزوع الأيونات وفق "
  "المعيار ANSI/ANS-16.1–2003 على مدى 90 يوماً. وحُلِّلت التغيّرات البنيوية بـ BET/BJH وXRD وATR-FTIR "
  "وTGA/DTG وSEM/EDS والرنين النووي المغناطيسي ²⁹Si. وأطال السيزيوم زمن الشكّ بنسبة 33% عند 2% و61% "
  "عند 5%، مما يدلّ على تباطؤ تفاعلات التكثيف، وعزّز تكوّن هلام ألومينوسيليكاتي غني بالألمنيوم يوحي "
  "بامتداد أكبر للتفاعل. وبعد خمسة أيام من الترشيح انخفض كسر الوحدات الغنية بالألمنيوم دون أثر معتبر "
  "في مقاومة الانضغاط. وكانت آلية الترشيح الأولية هي الانتشار، ثم انتقلت إلى الاستنزاف لاحقاً. وقد "
  "أتاح توظيف ²⁹Si MAS NMR مع منهجيات تحليلية أخرى توضيح أثر السيزيوم في بنية الجيوبوليمر وآليات "
  "تثبيته وتغيّراته البنيوية أثناء الترشيح. وتوحي القيم العالية لمعامل قابلية الترشيح (نحو 10) والأداء "
  "الميكانيكي الجيد بأن جيوبوليمرات الرماد المتطاير حلٌّ اقتصادي ومستدام لتخزين النفايات النووية "
  "الحاوية للسيزيوم بأمان.",
 "tian2019_espr":
  "أُعيد تدوير الرماد المتطاير الفحمي، بوصفه نفاية صلبة من محطات الطاقة الفحمية، لتخليق زيوليت A "
  "وجيوبوليمر استُخدما في تثبيت/تصليب ⁺Cs و²⁺Sr من المحاليل المائية. وقد توافقت بيانات الامتزاز مع "
  "النماذج الحركية والديناميكية الحرارية، واستُكشِفت تغيّرات البنية المجهرية للزيوليت A بعد تحميل "
  "الأيونات بـ XRD وFTIR ورامان وTG-DTA وأيزوثرم امتزاز/انفكاك النيتروجين. وأُجري تصليب الزيوليتات "
  "المستهلكة بالجيوبوليمر وقُيِّم. وتبيّن سيادة آلية الرتبة الثانية الزائفة، وأن الانتشار الغشائي — "
  "وفق معادلة بويد — يبدو حاكماً لعملية الامتزاز. وكانت السعتان العظميان بنموذج لانغموير 2.12 و1.93 "
  "ميلي مول/غ للسيزيوم والسترونتيوم على التوالي. وأثناء التبادل الأيوني مال ⁺Cs إلى شغل موضع الحلقة "
  "الثمانية الأعضاء، بينما مال ²⁺Sr إلى استبدال ⁺Na في الحلقة السداسية، مؤدّياً إلى تغيّرات مختلفة في "
  "بنية الزيوليت. كما يمكن أن يكون الجيوبوليمر مصفوفةً واعدةً لمعالجة النفايات المشعة لأن كسر الترشيح "
  "انخفض كثيراً بعد التصليب.",
 "vandevenne2018":
  "من أكبر التحديات أمام الصناعة النووية التثبيت الآمن والمستدام للنفايات المشعة (RAW). وتقوم المصفوفات "
  "الأكثر شيوعاً حالياً لتثبيت النفايات المنخفضة والمتوسطة المستوى على الإسمنت البورتلاندي العادي. "
  "أمّا النويدات الأصعب تثبيتاً مثل السيزيوم (⁺Cs) والسترونتيوم (²⁺Sr) فقد دُرست لها مصفوفات بديلة، من "
  "أوعدها المواد المنشَّطة قلوياً (AAM). غير أن اختلاف تراكيب السلائف واستخدام أنواع مختلفة من محاليل "
  "التنشيط يُصعّب فهم أثر تركيب السليفة في تثبيت النويدات المُدخَلة. لذلك طُوِّرت ست تركيبات من خبث "
  "‏Ca-Si-Al المخلَّق مخبرياً لتكون سلائف لموادّ منشَّطة قلوياً منخفضة القلوية بغية دراسة سلوك تثبيتها. "
  "وتحقّقت قدرات تثبيت بلغت 97.6% للسيزيوم و99.9% للسترونتيوم عند تحميل 1% وزناً بالترشيح 7 أيام عند "
  "20 °م في ماء Milli Q. ويكون تثبيت السيزيوم أعلى عند نسبتَي Si/Al وCa/(Si+Al) الأدنى، بينما يكون "
  "تثبيت السترونتيوم أعلى عند نسبة Ca/(Si+Al) الأدنى ومستقلاً عن نسبة Si/Al. وتقدّم النتائج فهماً "
  "أعمق لسلوك تثبيت هذه المواد وتشجّع مزيداً من البحث والتطبيق في تثبيت النفايات المشعة.",
 "varon2025":
  "رُكِّز في هذا المقال على تطوير مازّات مستقرة كيميائياً وعالية الأداء لإزالة ²⁺Sr انتقائياً من مياه "
  "الصرف النووية، بما يتيح تخزيناً وتخلصاً أكثر أماناً على المدى الطويل. وبفضل الشحنة السالبة "
  "للألمنيوم العامل موقعَ تبادل امتزازي (AlIV)، تعزّز الجيوبوليمرات التبادل الأيوني، مُبديةً حركية "
  "امتزاز وسعات تبادل مواتية. وطُوِّرت مقاربة جديدة لتوصيف الجيوبوليمرات بشمولية وإرساء علاقات "
  "البنية-الامتزاز. وتبحث الدراسة كيف تؤثّر عوامل تخليق رئيسة — تحديداً نسبتا Si/Al وH2O/M2O المولّيتان "
  "— في الخصائص البنيوية المُوصَّفة بامتزاز/انفكاك النيتروجين والرنين النووي المغناطيسي ²⁷Al و²⁹Si، "
  "اللذين يسبران البنية وتركيز AlIV لتقييم أثرهما في سلوك امتزاز السترونتيوم والكالسيوم. وتزيد زيادة "
  "نسبة Si/Al المساميةَ لكنها تخفض تركيز AlIV، فتقلّ سعة الامتزاز؛ ويُعزى هذا الانخفاض إلى تباعد "
  "أكبر بين وحدات AlIV وتحوّل بنيوي نحو شبكة أغلب على السيليكون. وطُوِّر نموذج سطحي ثنائي الأبعاد "
  "قائم على مراكز السيليكون Q4(mAl) من طيف ²⁹Si لربطه بخصائص الامتزاز، ولوحظت انتقائية للسترونتيوم "
  "عند Si/Al = 1.52.",
 "varon_leached2026":
  "الجيوبوليمرات موادّ واعدة للاستخدام مازّةً للسترونتيوم في عمليات الأعمدة الثابتة لإزالة تلوّث مياه "
  "الصرف النووية. ومع ذلك تؤثّر معاملات تركيب الجيوبوليمر — نسبتا SiO2/M2O وH2O/M2O المولّيتان — في "
  "تطوّر البنية المجهرية وخصائص امتزاز السترونتيوم تحت الترشيح المائي. فزيادة نسبة SiO2/M2O ترفع "
  "مقاومة الانضغاط (من 16 إلى 35 ميغاباسكال) وتقلّل ترشيح السيليكون في الماء (من 1.31 إلى 0.26 "
  "ميلي مول/غ)؛ وتنتج هذه المقاومة للترشيح من بنية أغنى بالسيليكون تعزّز مقاومة التحلّل المائي. أمّا "
  "زيادة نسبة H2O/M2O فتخفض المتانة (من 59 إلى 28 ميغاباسكال) وتزيد الترشيح لزيادة المسامية. ثم إن "
  "الترشيح المائي يخفض خصائص امتزاز الجيوبوليمرات بإطلاق AlIV التي تمثّل مواقع تبادل للكاتيونات. ومع "
  "ذلك حقّقت التركيبة الأدنى تركيزاً في AlIV أعلى معامل توزيع (K_D) وأسرع حركية، مبرزةً الدور الحاسم "
  "لبيئة السيليكون في احتجاز الكاتيونات على نحو أمثل، بما يفوق التركيز الكلّي لمواقع AlIV.",
 "walkley2020":
  "كثيراً ما تُثبَّت مجاري النفايات المشعة الحاوية لـ ⁹⁰Sr، الناجمة عن توليد الطاقة النووية وعمليات "
  "التنظيف البيئي، في الإسمنتات للحدّ من ترشيح النويدات. ولضعف توافق بعض النفايات مع الإسمنت "
  "البورتلاندي، تُدرَس بدائل مثل الجيوبوليمرات الألومينوسيليكاتية القلوية. ونُظهر هنا أن "
  "الجيوبوليمرات غير المنتظمة (هلامات (N,K)-A-S-H) المتكوّنة بالتنشيط القلوي للميتاكاولين تستوعب "
  "بسهولة كاتيونَي الأرض القلوية ²⁺Sr و²⁺Ca في شبكتها الألومينوسيليكاتية. والناتج الرئيس في الهلامات "
  "المعالَجة عند 20 °م و80 °م هو هلام (N,K)-A-S-H غني بالألمنيوم تام البلمرة يضمّ Al وSi في تناسق "
  "رباعي السطوح، مع Si في مواقع Q4(4Al) وQ4(3Al)، و⁺Na و⁺K يوازنان الشحنة السالبة الناتجة عن "
  "³⁺Al الرباعي التناسق. ويتكوّن الفوجاسيت-Na وزيوليت Na-A المستبدَل جزئياً بالسترونتيوم في الهلامات "
  "المعالَجة عند 80 °م. ويزيح إدخال ²⁺Sr أو ²⁺Ca بعض ⁺Na و⁺K من مواقع موازنة الشحنة، مع انخفاض طفيف "
  "في نسبة Si/Al. ويُحدث ²⁺Ca و²⁺Sr التغيّرات البنيوية نفسها جوهرياً في الهلامات.",
 "xiang2021":
  "يفيد فصل ⁷³¹Cs — ذي فترة الاضمحلال الطويلة والحرارة الاضمحلالية العالية — وإزالتُه في المعالجة "
  "اللاحقة للنفايات المشعة والتخلص منها. في هذه الدراسة حُضِّرت رغوات جيوبوليمرية (GFs) ذات أداء "
  "ميكانيكي مناسب وكثافة 300–600 كغ/م³ بطريقة الرغوة المُصنَّعة مسبقاً، واقتُرحت وحداتٍ مازّةً "
  "لـ ⁷³¹Cs استناداً إلى بنيتها الشبيهة بالزيوليت والمسام الدقيقة والمتوسطة الهرمية. وأبدت الرغوة "
  "الكتلية سعة امتزاز لافتة (192.14 ملغ/غ) وكفاءة انفكاك (64.92%) للسيزيوم المحاكى في المحلول المائي. "
  "واتّبعت عملية الامتزاز حركية الرتبة الثانية الزائفة ونموذج لانغموير، وكانت تلقائية وماصّة للحرارة "
  "ومصحوبة بازدياد الإنتروبيا. وأخيراً اقتُرحت آلية امتزاز السيزيوم المحاكى وتركيزه في الرغوات.",
}


def _vancouver(rec: dict) -> str:
    """Vancouver reference: Authors. Title. Journal. Year;Vol(Issue):Pages. doi:DOI"""
    au = rec.get("authors") or []
    if len(au) > 6:
        authors = ", ".join(au[:6]) + ", et al"
    else:
        authors = ", ".join(au)
    bits = [f"{authors}." if authors else "", f"{rec.get('title','').rstrip('.')}."]
    if rec.get("journal"):
        bits.append(f"{rec['journal']}.")
    ym = str(rec.get("year") or "")
    tail = ym
    if rec.get("volume"):
        tail += f";{rec['volume']}"
        if rec.get("issue"):
            tail += f"({rec['issue']})"
        if rec.get("page"):
            tail += f":{rec['page']}"
    if tail:
        bits.append(tail + ".")
    return " ".join(b for b in bits if b)


def load() -> list[dict]:
    cr = json.loads(SOURCES.read_text())
    out = []
    for doi, rec in cr.items():
        if "title" not in rec:
            continue
        lab = rec["label"]
        src = rec.get("abstract_source", "")
        abstract = rec.get("abstract", "")
        status = {
            "Crossref": "publisher abstract (verbatim, via Crossref)",
            "verified from PDF": "publisher abstract (verbatim, verified against the paper's PDF)",
        }.get(src, "not included — the paper is not held on disk or is a scanned image; "
                   "retrieve from the DOI")
        out.append({
            "label": lab,
            "title": rec.get("title", ""),
            "authors": "; ".join(rec.get("authors") or []),
            "journal": rec.get("journal", ""),
            "year": rec.get("year"),
            "volume": rec.get("volume") or "",
            "issue": rec.get("issue") or "",
            "pages": rec.get("page") or "",
            "doi": doi,
            "url": f"https://doi.org/{doi}",
            "vancouver": _vancouver(rec),
            "abstract_en": abstract,
            "abstract_status": status,
            "abstract_ar": ARABIC.get(lab, ""),
            "role": ROLE.get(lab, ""),
            "licence": "; ".join(rec.get("license") or []) or "not declared",
        })
    return sorted(out, key=lambda r: (r["year"] or 0, r["label"]))


def write_xlsx(rows: list[dict]) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Bibliography"
    cols = [("#", 5), ("Source key", 18), ("Title", 60), ("Authors", 40), ("Journal", 32),
            ("Year", 7), ("Vol", 7), ("Issue", 7), ("Pages", 12), ("DOI", 32), ("URL", 34),
            ("Abstract (original, verbatim)", 80), ("Abstract status", 30),
            ("الملخص بالعربية", 80), ("Role in this study", 46), ("Licence (Crossref)", 26)]
    head = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="0D5C63")
    for i, (name, width) in enumerate(cols, 1):
        c = ws.cell(row=1, column=i, value=name)
        c.font = head
        c.fill = fill
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"

    for r, row in enumerate(rows, start=2):
        vals = [r - 1, row["label"], row["title"], row["authors"], row["journal"], row["year"],
                row["volume"], row["issue"], row["pages"], row["doi"], row["url"],
                row["abstract_en"], row["abstract_status"], row["abstract_ar"], row["role"],
                row["licence"]]
        for i, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=i, value=v)
            c.alignment = Alignment(wrap_text=True, vertical="top")
        # Arabic column reads right-to-left
        ws.cell(row=r, column=14).alignment = Alignment(
            wrap_text=True, vertical="top", horizontal="right", readingOrder=2)

    # a second sheet documenting the method, so the file explains itself
    ws2 = wb.create_sheet("About")
    about = [
        ["GEOMIND-R — corpus bibliography"],
        [""],
        ["Every source publication this study drew on, with metadata resolved from Crossref by DOI."],
        [""],
        ["Abstracts"],
        ["Included ONLY where the publisher supplies them through Crossref, so every word in the"],
        ["'Abstract (original, verbatim)' column is attributable. Automated extraction from PDF"],
        ["text was trialled and REJECTED: it produced a confirmed wrong-paper match (a slag/Portland"],
        ["cement abstract attached to a zeolite ion-exchange paper). Unverified text is not shipped."],
        ["Rows without a publisher abstract carry the DOI, which resolves to the authoritative version."],
        [""],
        ["Arabic"],
        ["Translations of the publisher abstracts, prepared for this compilation. The English column"],
        ["remains the authoritative verbatim text."],
        [""],
        ["Copyright"],
        ["Publisher abstracts are the copyright of their publishers. This file is a working"],
        ["bibliography for the authors' own submission and is NOT part of the public data deposit."],
        [""],
        ["Cite the dataset: Fetna M, Hammal A. GEOMIND-R. https://doi.org/10.5281/zenodo.21510123"],
    ]
    for i, line in enumerate(about, 1):
        ws2.cell(row=i, column=1, value=line[0])
    ws2.column_dimensions["A"].width = 100
    ws2.cell(row=1, column=1).font = Font(bold=True, size=13)

    out = _HERE / "GEOMIND-R-bibliography.xlsx"
    wb.save(out)
    return out


def write_docx(rows: list[dict]) -> Path:
    md = ["# GEOMIND-R — corpus bibliography (Vancouver style)", "",
          "Every source publication this study drew on. Metadata resolved from Crossref by DOI.",
          "Abstracts are reproduced **only** where the publisher supplies them, so that every",
          "quoted word is verbatim and attributable; the remaining entries carry the DOI, which",
          "resolves to the authoritative version. Arabic translations follow each abstract.", "",
          "*Publisher abstracts remain the copyright of their publishers. This is a working",
          "bibliography for the authors' submission, not part of the public data deposit.*", "",
          "---", ""]
    for i, r in enumerate(rows, 1):
        md.append(f"**{i}.** {r['vancouver']} doi:{r['doi']}")
        md.append("")
        if r["role"]:
            md.append(f"*Role in this study:* {r['role']}")
            md.append("")
        if r["abstract_en"]:
            md.append(f"**Abstract.** {r['abstract_en']}")
            md.append("")
            if r["abstract_ar"]:
                md.append(f"**الملخص.** {r['abstract_ar']}")
                md.append("")
        else:
            md.append(f"*Abstract:* {r['abstract_status']} — <{r['url']}>")
            md.append("")
        md.append("")
    src = _HERE / "GEOMIND-R-bibliography.md"
    src.write_text("\n".join(md))
    out = _HERE / "GEOMIND-R-bibliography.docx"
    subprocess.run(["pandoc", str(src), "-o", str(out), "--standalone"], check=True)
    return out


def main() -> None:  # pragma: no cover
    rows = load()
    x = write_xlsx(rows)
    d = write_docx(rows)
    withabs = sum(1 for r in rows if r["abstract_en"])
    print(f"wrote {x}\nwrote {d}\n{len(rows)} references, {withabs} with a publisher abstract "
          f"({sum(1 for r in rows if r['abstract_ar'])} translated)")


if __name__ == "__main__":  # pragma: no cover
    main()
